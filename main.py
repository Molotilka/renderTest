from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

app = FastAPI()

@app.post("/export")
async def export_excel(request: Request):
    data = await request.json()
    # Meta - Comp, Manag ...
    meta = {}
    table_data = []
    for item in data:
        if len(item) <= 2 and not all(k in item for k in ["Category", "Brand", "Model", "Name", "Available QTY", "DDP (RUB)"]):
            meta.update(item)
        else:
            table_data.append(item)

    wb = Workbook()
    ws = wb.active
    ws.title = "Shavers"

    # Верхняя часть
    ws["A1"] = "Company:"
    ws["B1"] = meta["Company Name"]
    ws["A2"] = "Manager:"
    ws["B2"] = meta["Full Name"]
    ws["A3"] = "E-mail:"
    ws["B3"] = meta["Mail"]
    ws["A4"] = "Date:"
    ws["B4"] = meta["Created"]

    for row in range(1, 5):
        ws[f"A{row}"].alignment = Alignment(horizontal="right")
        ws[f"A{row}"].font = Font(color="C0C0C0")
        ws[f"B{row}"].font = Font(color="808080", bold=True)

    # Заголовки
    headers = ["Category", "Brand", "Model", "Name", "Available QTY", "DDP (RUB)"]
    ws.append([])
    ws.append(headers)

    # Данные
    data = []

    for row in table_data:
        data.append([
            row.get("Category", ""),
            row.get("Brand", ""),
            row.get("Model", ""),
            row.get("Name", ""),
            row.get("Available QTY", ""),
            row.get("DDP (RUB)", "")
        ])

    count = 7
    for row in data:
        ws.append(row)
        ws.cell(row=count, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=count, column=6).alignment = Alignment(horizontal="center")
        count += 1

    # Стили заголовков
    for cell in ws[6]:
        cell.border = Border(bottom=Side(style='thick', color='000000'))
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="98FB98")

    # Ширина колонок
    column_widths = [25, 25, 15, 50, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Сохраняем в память
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=shavers_report.xlsx"}
    )
